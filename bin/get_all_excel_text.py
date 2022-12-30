import nltk

nltk.download('words')

from openpyxl import load_workbook
from nltk.corpus import words

import re


def is_word(value):
    # Use a regular expression to match only letters
    return str(bool(re.match(r'^[a-zA-Z]+$', value)))


def get_all_excel_text(path):
    # Load the workbook
    workbook = load_workbook(path)
    # Get the worksheets
    worksheets = workbook.worksheets

    english_words = []
    # Iterate through the worksheets
    for sheet in worksheets:
        # Iterate through the rows
        for row in sheet.iter_rows():
            # Iterate through the cells
            for cell in row:
                # Get the value of the cell
                value = cell.value
                # print(value)
                # Check if the value is an English word
                if not value == None and not value == "" and not value=="nan":
                    if is_word(str(value)):
                        # Add the word to a list
                        try:
                            english_words.append(value.lower())
                        except:
                            pass
    return english_words
