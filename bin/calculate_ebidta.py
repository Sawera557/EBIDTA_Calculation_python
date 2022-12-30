import openpyxl
import re


def extract_money(text):
    money_pattern = r'\$?\s*[\d,]+(?:\.\d+)?'
    text = str(text)

    money_numbers = re.findall(money_pattern, text)
    if not money_numbers==[]:
        if "," in money_numbers[0]:
            money_numbers = (money_numbers[0]).replace(",", "")
        elif " " in money_numbers[0]:
            money_numbers = (money_numbers[0]).replace(" ", "")
    return money_numbers


def get_values_from_files(finance_files):

    ebt_values = []

    for f in finance_files:
        file_path = f["file_path"]
        filename = f["filename"]
        book = openpyxl.load_workbook(file_path)
        sheet = book.worksheets[0]
        max_row = sheet.max_row + 1
        max_col = sheet.max_column +1
        #filename = "cashflow"

        if filename == "cashflow":

            net_value = ""
            dep_amor = ""

            for i in range(1, max_row):
                for j in range(1, max_col):  # corrected to range(1, max_col)
                    value = str(sheet.cell(row=i, column=j).value)
                    if not value == None and not "none" in value:
                        value = value.lower()

                        if "net income" == value:
                            found = False
                            col = j + 1
                            for co in range(col, max_col):
                                temp_net_val = sheet.cell(row=i, column=co).value
                                check_net_val = extract_money(temp_net_val)
                                if not check_net_val == "" and not check_net_val == []:
                                    net_value = float(check_net_val[0])

                                    found=True
                                if found==True:
                                    break

                        if "depreciation and amortization" in value or "depreciation & amortization" in value:
                            found = False
                            col = j + 1
                            for co in range(col, max_col):
                                temp_net_val = sheet.cell(row=i, column=co).value
                                check_net_val = extract_money(temp_net_val)
                                if not check_net_val == "" and not check_net_val == []:
                                    dep_amor = float(check_net_val[0])

                                    found = True
                                    if found == True:
                                        break



        elif filename=="income":
            tot_expenxe = ""
            for i in range(1, max_row):
                for j in range(1, max_col):
                    value = str(sheet.cell(row=i, column=j).value)

                    if not value == None and not value == "none":
                        value = value.lower()
                        if "total expenses" in value:
                            found = False
                            col = j + 1
                            for co in range(col, max_col):
                                temp_net_val = sheet.cell(row=i, column=co).value
                                check_net_val = extract_money(temp_net_val)
                                if not check_net_val=="" and not check_net_val==[]:
                                    tot_expenxe = float(check_net_val[0])
                                    found = True
                                if found==True:
                                    break




        elif filename=="balance":
            in_tax = ""

            for i in range(1, max_row):
                for j in (1, max_col):
                    value = str(sheet.cell(row=i,column=j).value)
                    if not value == None and not value=="none":
                        value = value.lower()

                        if "income taxes payable" in value:
                            found = False
                            col = j+1
                            for co in range(col, max_col):
                                temp_net_val = sheet.cell(row=i, column=co).value
                                check_net_val = extract_money(temp_net_val)
                                if not check_net_val=="" and not check_net_val==[]:
                                    in_tax = float(check_net_val[0])
                                    found = True
                                if found == True:
                                    break

    ebt_values.append({"net_value": net_value, "dep_amor": dep_amor, "tot_expenxe": tot_expenxe, "in_tax": in_tax})
    return ebt_values
